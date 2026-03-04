#!/usr/bin/env python3
"""
export_hospitals.py
將 hospitals.json 匯出為 Excel 檔案（hospitals.xlsx）

用法: python export_hospitals.py
"""
import sys
import json
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

HOSPITALS_JSON = Path("src/data/hospitals.json")
OUTPUT_FILE    = Path("hospitals.xlsx")

# 欄位定義：(欄位名稱, JSON 鍵, 欄寬)
COLUMNS = [
    ("機構代碼",   "id",             14),
    ("醫院名稱",   "name",           36),
    ("縣市",       "city",            8),
    ("行政區",     "district",       10),
    ("地址",       "address",        40),
    ("電話",       "phone",          16),
    ("官方網站",   "website",        40),
    ("網路掛號",   "appointmentUrl", 40),
    ("科別",       "services",       30),
]

# 資料列樣式
DATA_FONT    = Font(name="微軟正黑體", size=10)
DATA_ALIGN   = Alignment(vertical="center", wrap_text=False)
ALT_FILL     = PatternFill("solid", fgColor="EFF6FF")   # 淺藍（偶數列）

# 邊框
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def apply_header_style(ws) -> None:
    """套用標頭列樣式（藍底白字 + 邊框）"""
    for cell in ws[1]:
        cell.fill      = PatternFill("solid", fgColor="2563EB")
        cell.font      = Font(name="微軟正黑體", bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border    = BORDER


def main():
    with open(HOSPITALS_JSON, encoding="utf-8") as f:
        hospitals = json.load(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "醫院清單"
    ws.freeze_panes = "A2"   # 凍結標頭列

    # ── 標頭 ──────────────────────────────────────────────────────────
    ws.append([col[0] for col in COLUMNS])
    apply_header_style(ws)

    # ── 資料列 ────────────────────────────────────────────────────────
    for row_idx, h in enumerate(hospitals, start=2):
        row_data = []
        for _, key, _ in COLUMNS:
            if key == "services":
                val = "、".join(h.get("services", []))
            else:
                val = h.get(key) or ""
            row_data.append(val)
        ws.append(row_data)

        # 套用樣式
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            cell.font      = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border    = BORDER
            if fill:
                cell.fill  = fill

            # 官網 / 掛號欄位加超連結
            col_key = COLUMNS[col_idx - 1][1]
            if col_key in ("website", "appointmentUrl") and cell.value:
                is_odd_row = row_idx % 2 == 1
                cell.hyperlink = cell.value
                cell.font = Font(
                    name="微軟正黑體", size=10,
                    color="2563EB", underline="single",
                    bold=is_odd_row,
                )

    # ── 欄寬 ─────────────────────────────────────────────────────────
    for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── 列高 ─────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22   # 標頭列
    for row_idx in range(2, len(hospitals) + 2):
        ws.row_dimensions[row_idx].height = 16

    # ── 自動篩選 ──────────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    # ── 統計工作表 ────────────────────────────────────────────────────
    ws_stats = wb.create_sheet("統計")
    ws_stats.column_dimensions["A"].width = 20
    ws_stats.column_dimensions["B"].width = 12

    ws_stats.append(["項目", "數量"])
    apply_header_style(ws_stats)

    total    = len(hospitals)
    has_web  = sum(1 for h in hospitals if h.get("website"))
    has_appt = sum(1 for h in hospitals if h.get("appointmentUrl"))
    no_web   = total - has_web
    no_appt  = total - has_appt

    city_counts = Counter(h["city"] for h in hospitals)

    stats_rows = [
        ("醫院總數",         total),
        ("有官方網站",       has_web),
        ("無官方網站",       no_web),
        ("有網路掛號連結",   has_appt),
        ("無網路掛號連結",   no_appt),
        ("",                 ""),
        ("── 各縣市醫院數 ──", ""),
    ] + sorted(city_counts.items(), key=lambda x: -x[1])

    for row in stats_rows:
        ws_stats.append(list(row))

    for row_idx in range(2, ws_stats.max_row + 1):
        for cell in ws_stats[row_idx]:
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN

    wb.save(OUTPUT_FILE)
    print(f"✓ 匯出完成：{OUTPUT_FILE}")
    print(f"  共 {total} 間醫院 | 有官網 {has_web} 間 | 有掛號連結 {has_appt} 間")


if __name__ == "__main__":
    main()
